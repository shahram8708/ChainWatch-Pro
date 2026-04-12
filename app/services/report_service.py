"""Report generation service for PDF and Excel exports."""

from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timedelta
from statistics import mean

from flask import current_app, render_template_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func
from weasyprint import CSS, HTML

from app.models.alert import Alert
from app.models.carrier import Carrier
from app.models.carrier_performance import CarrierPerformance
from app.models.disruption_score import DisruptionScore
from app.models.organisation import Organisation
from app.models.route_recommendation import RouteRecommendation
from app.models.shipment import Shipment

logger = logging.getLogger(__name__)


REPORT_TYPES = {
    "monthly_performance": {
        "name": "Monthly Performance Report",
        "description": (
            "A complete operational performance snapshot across shipments, delivery reliability, and disruptions. "
            "It includes rerouting impact, carrier outcomes, and week-by-week service trend indicators for leadership."
        ),
        "available_formats": ["pdf", "excel"],
        "default_date_range_days": 30,
    },
    "carrier_comparison": {
        "name": "Carrier Comparison Report",
        "description": (
            "A benchmark report comparing all active carriers by OTD, delay patterns, and reliability quality. "
            "It highlights trend direction and lane-level strengths and weaknesses for allocation strategy."
        ),
        "available_formats": ["pdf", "excel"],
        "default_date_range_days": 90,
    },
    "lane_risk_analysis": {
        "name": "Lane Risk Analysis Report",
        "description": (
            "A route-by-route risk profile of disruption probability, delay impact, and cargo exposure concentration. "
            "It surfaces the highest-risk lanes first to support mitigation planning and routing policy updates."
        ),
        "available_formats": ["pdf", "excel"],
        "default_date_range_days": 90,
    },
    "disruption_audit": {
        "name": "Disruption Audit Report",
        "description": (
            "A forensic audit of disruption trajectories per shipment, including DRS band crossings and interventions. "
            "It provides traceability for compliance, post-mortems, and operational governance reviews."
        ),
        "available_formats": ["pdf", "excel"],
        "default_date_range_days": 30,
    },
}


def _to_float(value, default=0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value, default=0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _date_key(year: int, month: int) -> int:
    return (int(year) * 100) + int(month)


def _safe_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return None


def _period_bounds(start_date: datetime, end_date: datetime) -> tuple[datetime, datetime]:
    start = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    return start, end


def _format_inr_lakh_crore(value) -> str:
    amount = _to_float(value, 0.0)
    abs_amount = abs(amount)

    if abs_amount >= 10000000:
        display = amount / 10000000
        return f"₹ {display:.2f} Cr"
    if abs_amount >= 100000:
        display = amount / 100000
        return f"₹ {display:.2f} L"
    return f"₹ {amount:,.2f}"


def _outcome_label(shipment) -> str:
    if shipment.actual_arrival and shipment.estimated_arrival:
        return "On-Time" if shipment.actual_arrival <= shipment.estimated_arrival else "Late"
    if shipment.status == "delivered":
        return "Delivered"
    return "In Progress"


def _month_start(dt: datetime) -> datetime:
    return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def _build_monthly_kpis(organisation_id, start_date, end_date, db_session) -> dict:
    data = _generate_monthly_performance_data(
        organisation_id,
        start_date,
        end_date,
        db_session,
        include_kpi_trends=False,
    )
    return {
        "total_shipments": data["total_shipments"],
        "otd_rate": data["otd_rate"],
        "average_drs": data["average_drs"],
        "critical_alerts": data["critical_alerts_generated"],
        "estimated_savings_inr": data["estimated_savings_inr"],
    }


def _calculate_kpi_trends(organisation_id, start_date, end_date, db_session) -> dict:
    period_days = max((end_date.date() - start_date.date()).days + 1, 1)
    prev_end = start_date - timedelta(days=1)
    prev_start = prev_end - timedelta(days=period_days - 1)

    current_kpis = _build_monthly_kpis(organisation_id, start_date, end_date, db_session)
    previous_kpis = _build_monthly_kpis(organisation_id, prev_start, prev_end, db_session)

    trends = {}
    for key, current_value in current_kpis.items():
        previous_value = _to_float(previous_kpis.get(key), 0.0)
        current_numeric = _to_float(current_value, 0.0)

        if previous_value == 0:
            delta_pct = 100.0 if current_numeric > 0 else 0.0
        else:
            delta_pct = ((current_numeric - previous_value) / abs(previous_value)) * 100.0

        direction = "neutral"
        if delta_pct > 0.15:
            direction = "up"
        elif delta_pct < -0.15:
            direction = "down"

        trends[key] = {
            "delta_pct": round(delta_pct, 2),
            "direction": direction,
        }

    return trends


def generate_report(report_type, organisation_id, start_date, end_date, output_format, db_session, app_context):
    """Generate and persist a report file for asynchronous export workflows."""

    try:
        app = app_context or current_app._get_current_object()
        start_dt, end_dt = _period_bounds(start_date, end_date)

        if report_type not in REPORT_TYPES:
            raise ValueError(f"Unknown report_type: {report_type}")
        if output_format not in {"pdf", "excel"}:
            raise ValueError(f"Unsupported output format: {output_format}")

        organisation = db_session.query(Organisation).filter(Organisation.id == organisation_id).first()
        if organisation is None:
            raise ValueError("Organisation not found")

        data_generators = {
            "monthly_performance": _generate_monthly_performance_data,
            "carrier_comparison": _generate_carrier_comparison_data,
            "lane_risk_analysis": _generate_lane_risk_data,
            "disruption_audit": _generate_disruption_audit_data,
        }

        pdf_generators = {
            "monthly_performance": generate_monthly_performance_pdf,
            "carrier_comparison": generate_carrier_comparison_pdf,
            "lane_risk_analysis": generate_lane_risk_pdf,
            "disruption_audit": generate_disruption_audit_pdf,
        }

        excel_generators = {
            "monthly_performance": generate_monthly_performance_excel,
            "carrier_comparison": generate_carrier_comparison_excel,
            "lane_risk_analysis": generate_lane_risk_excel,
            "disruption_audit": generate_disruption_audit_excel,
        }

        data = data_generators[report_type](organisation_id, start_dt, end_dt, db_session)

        if output_format == "pdf":
            file_bytes = pdf_generators[report_type](data, organisation, start_dt, end_dt, app)
            extension = "pdf"
        else:
            file_bytes = excel_generators[report_type](data, organisation, start_dt, end_dt)
            extension = "xlsx"

        output_dir = app.config.get(
            "REPORT_OUTPUT_DIR",
            os.path.join(app.root_path, "..", "static", "reports"),
        )
        os.makedirs(output_dir, exist_ok=True)

        org_id_short = str(organisation_id).replace("-", "")[:8]
        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = (
            f"{report_type}_{org_id_short}_{start_dt.strftime('%Y%m%d')}_"
            f"{end_dt.strftime('%Y%m%d')}_{timestamp}.{extension}"
        )
        file_path = os.path.join(output_dir, filename)

        with open(file_path, "wb") as output_file:
            output_file.write(file_bytes)

        return {
            "success": True,
            "file_path": file_path,
            "filename": filename,
            "error": None,
        }
    except Exception as exc:
        logger.exception("Report generation failed type=%s organisation_id=%s", report_type, organisation_id)
        return {
            "success": False,
            "file_path": None,
            "filename": None,
            "error": str(exc),
        }


def _generate_monthly_performance_data(
    organisation_id,
    start_date,
    end_date,
    db_session,
    include_kpi_trends: bool = True,
):
    """Collect all monthly performance metrics and tabular datasets."""

    period_shipments = (
        db_session.query(Shipment)
        .filter(
            Shipment.organisation_id == organisation_id,
            Shipment.estimated_departure >= start_date,
            Shipment.estimated_departure <= end_date,
            Shipment.is_archived.is_(False),
        )
        .all()
    )

    shipment_ids = [shipment.id for shipment in period_shipments]

    completed_shipments = [
        shipment
        for shipment in period_shipments
        if shipment.actual_arrival is not None and start_date <= shipment.actual_arrival <= end_date
    ]

    on_time_shipments = [
        shipment
        for shipment in completed_shipments
        if shipment.estimated_arrival is not None and shipment.actual_arrival <= shipment.estimated_arrival
    ]
    delayed_shipments = [
        shipment
        for shipment in completed_shipments
        if shipment.estimated_arrival is not None and shipment.actual_arrival > shipment.estimated_arrival
    ]

    total_shipments = len(period_shipments)
    completed_count = len(completed_shipments)
    on_time_count = len(on_time_shipments)
    delayed_count = len(delayed_shipments)
    otd_rate = round((on_time_count / completed_count) if completed_count else 0.0, 4)

    avg_drs = (
        db_session.query(func.coalesce(func.avg(DisruptionScore.drs_total), 0.0))
        .join(Shipment, Shipment.id == DisruptionScore.shipment_id)
        .filter(
            Shipment.organisation_id == organisation_id,
            DisruptionScore.computed_at >= start_date,
            DisruptionScore.computed_at <= end_date,
        )
        .scalar()
    )

    cargo_value_at_risk = (
        db_session.query(func.coalesce(func.sum(Shipment.cargo_value_inr), 0.0))
        .filter(
            Shipment.organisation_id == organisation_id,
            Shipment.estimated_departure >= start_date,
            Shipment.estimated_departure <= end_date,
            Shipment.disruption_risk_score > 60,
            Shipment.is_archived.is_(False),
        )
        .scalar()
    )

    critical_alerts_generated = (
        db_session.query(Alert)
        .filter(
            Alert.organisation_id == organisation_id,
            Alert.severity == "critical",
            Alert.created_at >= start_date,
            Alert.created_at <= end_date,
        )
        .count()
    )

    acknowledged_alert_rows = (
        db_session.query(Alert)
        .filter(
            Alert.organisation_id == organisation_id,
            Alert.is_acknowledged.is_(True),
            Alert.acknowledged_at.isnot(None),
            Alert.acknowledged_at >= start_date,
            Alert.acknowledged_at <= end_date,
        )
        .all()
    )

    resolution_hours = []
    for alert in acknowledged_alert_rows:
        if alert.created_at and alert.acknowledged_at:
            resolution_hours.append((alert.acknowledged_at - alert.created_at).total_seconds() / 3600.0)

    avg_resolution_hours = round(mean(resolution_hours), 2) if resolution_hours else 0.0

    rerouting_rows = (
        db_session.query(RouteRecommendation, Shipment)
        .join(Shipment, Shipment.id == RouteRecommendation.shipment_id)
        .filter(
            Shipment.organisation_id == organisation_id,
            RouteRecommendation.decided_at.isnot(None),
            RouteRecommendation.decided_at >= start_date,
            RouteRecommendation.decided_at <= end_date,
            RouteRecommendation.status.in_(["approved", "dismissed"]),
        )
        .order_by(RouteRecommendation.decided_at.desc())
        .all()
    )

    approved_count = len([item for item, _ in rerouting_rows if item.status == "approved"])
    dismissed_count = len([item for item, _ in rerouting_rows if item.status == "dismissed"])
    approved_cost_delta_sum = round(
        sum(_to_float(item.cost_delta_inr, 0.0) for item, _ in rerouting_rows if item.status == "approved"),
        2,
    )
    estimated_savings_inr = round(
        sum(max(0.0, -_to_float(item.cost_delta_inr, 0.0)) for item, _ in rerouting_rows if item.status == "approved"),
        2,
    )

    rerouting_decisions = []
    for recommendation, shipment in rerouting_rows:
        rerouting_decisions.append(
            {
                "date": recommendation.decided_at,
                "shipment_reference": shipment.external_reference or str(shipment.id),
                "option_chosen": recommendation.option_label,
                "cost_delta_inr": _to_float(recommendation.cost_delta_inr, 0.0),
                "revised_eta": recommendation.revised_eta,
                "outcome": _outcome_label(shipment),
                "status": recommendation.status,
            }
        )

    top_carrier_rows = (
        db_session.query(
            Carrier.id.label("carrier_id"),
            Carrier.name.label("carrier_name"),
            func.count(Shipment.id).label("shipment_count"),
            func.sum(
                case(
                    (
                        func.coalesce(Shipment.actual_arrival <= Shipment.estimated_arrival, False),
                        1,
                    ),
                    else_=0,
                )
            ).label("on_time_count"),
            func.sum(
                case(
                    (Shipment.actual_arrival.isnot(None), 1),
                    else_=0,
                )
            ).label("completed_count"),
        )
        .join(Shipment, Shipment.carrier_id == Carrier.id)
        .filter(
            Shipment.organisation_id == organisation_id,
            Shipment.estimated_departure >= start_date,
            Shipment.estimated_departure <= end_date,
            Shipment.is_archived.is_(False),
        )
        .group_by(Carrier.id, Carrier.name)
        .order_by(func.count(Shipment.id).desc())
        .limit(5)
        .all()
    )

    top_carriers = []
    for row in top_carrier_rows:
        completed = _to_int(row.completed_count, 0)
        otd = (_to_float(row.on_time_count, 0.0) / completed) if completed else 0.0
        top_carriers.append(
            {
                "carrier_id": str(row.carrier_id),
                "carrier_name": row.carrier_name,
                "shipment_count": _to_int(row.shipment_count, 0),
                "otd_rate": round(otd, 4),
            }
        )

    shipment_volume_by_mode_rows = (
        db_session.query(
            Shipment.mode,
            func.count(Shipment.id).label("count"),
        )
        .filter(
            Shipment.organisation_id == organisation_id,
            Shipment.estimated_departure >= start_date,
            Shipment.estimated_departure <= end_date,
            Shipment.is_archived.is_(False),
        )
        .group_by(Shipment.mode)
        .order_by(func.count(Shipment.id).desc())
        .all()
    )

    shipment_volume_by_mode = [
        {
            "mode": row.mode,
            "count": _to_int(row.count, 0),
        }
        for row in shipment_volume_by_mode_rows
    ]

    week_group = {}
    for shipment in completed_shipments:
        if shipment.actual_arrival is None:
            continue
        iso = shipment.actual_arrival.isocalendar()
        key = f"{iso.year}-W{iso.week:02d}"
        bucket = week_group.setdefault(key, {"completed": 0, "on_time": 0})
        bucket["completed"] += 1
        if shipment.estimated_arrival and shipment.actual_arrival <= shipment.estimated_arrival:
            bucket["on_time"] += 1

    week_otd_trend = []
    for key in sorted(week_group.keys()):
        completed = week_group[key]["completed"]
        on_time = week_group[key]["on_time"]
        week_otd_trend.append(
            {
                "iso_week": key,
                "completed": completed,
                "on_time": on_time,
                "otd_rate": round((on_time / completed) if completed else 0.0, 4),
            }
        )

    peak_drs_map = {}
    if shipment_ids:
        peak_rows = (
            db_session.query(
                DisruptionScore.shipment_id,
                func.max(DisruptionScore.drs_total).label("peak_drs"),
            )
            .filter(DisruptionScore.shipment_id.in_(shipment_ids))
            .group_by(DisruptionScore.shipment_id)
            .all()
        )
        peak_drs_map = {
            row.shipment_id: round(_to_float(row.peak_drs, 0.0), 2)
            for row in peak_rows
        }

    shipment_detail = []
    for shipment in period_shipments:
        delay_hours = None
        if shipment.actual_arrival and shipment.estimated_arrival:
            delay_hours = round((shipment.actual_arrival - shipment.estimated_arrival).total_seconds() / 3600.0, 2)

        shipment_detail.append(
            {
                "shipment_id": str(shipment.id),
                "external_reference": shipment.external_reference or "-",
                "carrier": shipment.carrier.name if shipment.carrier else "Unassigned",
                "mode": shipment.mode,
                "origin": shipment.origin_port_code,
                "destination": shipment.destination_port_code,
                "status": shipment.status,
                "estimated_departure": shipment.estimated_departure,
                "estimated_arrival": shipment.estimated_arrival,
                "actual_arrival": shipment.actual_arrival,
                "delay_hours": delay_hours,
                "peak_drs": peak_drs_map.get(shipment.id, _to_float(shipment.disruption_risk_score, 0.0)),
                "cargo_value_inr": _to_float(shipment.cargo_value_inr, 0.0),
            }
        )

    carrier_performance_data = []
    grouped_carrier = {}
    for shipment in period_shipments:
        carrier_name = shipment.carrier.name if shipment.carrier else "Unassigned"
        key = (carrier_name, shipment.mode)
        bucket = grouped_carrier.setdefault(
            key,
            {
                "carrier_name": carrier_name,
                "mode": shipment.mode,
                "total_shipments": 0,
                "completed": 0,
                "on_time": 0,
                "delay_hours": [],
                "crs_score": 0.0,
            },
        )
        bucket["total_shipments"] += 1
        if shipment.actual_arrival and shipment.estimated_arrival:
            bucket["completed"] += 1
            if shipment.actual_arrival <= shipment.estimated_arrival:
                bucket["on_time"] += 1
            else:
                bucket["delay_hours"].append(
                    max((shipment.actual_arrival - shipment.estimated_arrival).total_seconds() / 3600.0, 0.0)
                )
        bucket["crs_score"] += max(0.0, 100.0 - _to_float(shipment.disruption_risk_score, 0.0))

    for _, bucket in grouped_carrier.items():
        completed = bucket["completed"]
        total = bucket["total_shipments"]
        avg_delay = mean(bucket["delay_hours"]) if bucket["delay_hours"] else 0.0
        crs_score = bucket["crs_score"] / total if total else 0.0
        carrier_performance_data.append(
            {
                "carrier_name": bucket["carrier_name"],
                "mode": bucket["mode"],
                "otd_rate": round((bucket["on_time"] / completed) if completed else 0.0, 4),
                "avg_delay_hours": round(avg_delay, 2),
                "total_shipments": total,
                "crs_score": round(crs_score, 2),
            }
        )

    carrier_performance_data.sort(key=lambda item: item["otd_rate"], reverse=True)

    alert_rows = (
        db_session.query(Alert, Shipment)
        .outerjoin(Shipment, Shipment.id == Alert.shipment_id)
        .filter(
            Alert.organisation_id == organisation_id,
            Alert.created_at >= start_date,
            Alert.created_at <= end_date,
        )
        .order_by(Alert.created_at.desc())
        .all()
    )

    alerts_data = []
    for alert, shipment in alert_rows:
        resolution = None
        if alert.is_acknowledged and alert.acknowledged_at and alert.created_at:
            resolution = round((alert.acknowledged_at - alert.created_at).total_seconds() / 3600.0, 2)
        alerts_data.append(
            {
                "alert_type": alert.alert_type,
                "severity": alert.severity,
                "title": alert.title,
                "shipment_reference": shipment.external_reference if shipment else "-",
                "created_at": alert.created_at,
                "acknowledged_at": alert.acknowledged_at,
                "resolution_hours": resolution,
            }
        )

    kpi_trends = {}
    if include_kpi_trends:
        kpi_trends = _calculate_kpi_trends(organisation_id, start_date, end_date, db_session)

    return {
        "total_shipments": total_shipments,
        "completed_shipments": completed_count,
        "on_time_shipments": on_time_count,
        "delayed_shipments": delayed_count,
        "otd_rate": round(otd_rate, 4),
        "average_drs": round(_to_float(avg_drs, 0.0), 2),
        "cargo_value_at_risk_inr": round(_to_float(cargo_value_at_risk, 0.0), 2),
        "critical_alerts_generated": int(critical_alerts_generated or 0),
        "alerts_acknowledged": len(acknowledged_alert_rows),
        "avg_alert_resolution_hours": avg_resolution_hours,
        "rerouting_approved_count": approved_count,
        "rerouting_dismissed_count": dismissed_count,
        "rerouting_cost_delta_inr": approved_cost_delta_sum,
        "estimated_savings_inr": estimated_savings_inr,
        "top_carriers": top_carriers,
        "shipment_volume_by_mode": shipment_volume_by_mode,
        "week_otd_trend": week_otd_trend,
        "rerouting_decisions": rerouting_decisions,
        "shipment_detail": shipment_detail,
        "carrier_performance": carrier_performance_data,
        "alerts": alerts_data,
        "kpi_trends": kpi_trends,
    }


def _generate_carrier_comparison_data(organisation_id, start_date, end_date, db_session):
    """Build carrier comparison metrics and trend signals for report export."""

    start_key = _date_key(start_date.year, start_date.month)
    end_key = _date_key(end_date.year, end_date.month)
    midpoint = start_date + (end_date - start_date) / 2

    perf_rows = (
        db_session.query(CarrierPerformance, Carrier)
        .join(Carrier, Carrier.id == CarrierPerformance.carrier_id)
        .filter(
            CarrierPerformance.organisation_id == organisation_id,
            ((CarrierPerformance.period_year * 100) + CarrierPerformance.period_month) >= start_key,
            ((CarrierPerformance.period_year * 100) + CarrierPerformance.period_month) <= end_key,
        )
        .all()
    )

    grouped = {}
    for perf, carrier in perf_rows:
        bucket = grouped.setdefault(
            carrier.id,
            {
                "carrier_id": str(carrier.id),
                "carrier_name": carrier.name,
                "mode_counts": {},
                "total_shipments": 0,
                "on_time": 0,
                "weighted_delay_sum": 0.0,
                "weighted_crs_sum": 0.0,
                "lanes": [],
                "current_total": 0,
                "current_on_time": 0,
                "prior_total": 0,
                "prior_on_time": 0,
            },
        )

        total = _to_int(perf.total_shipments, 0)
        on_time = _to_int(perf.on_time_count, 0)
        avg_delay = _to_float(perf.avg_delay_hours, 0.0)
        crs = _to_float(perf.reliability_score, 0.0)

        bucket["total_shipments"] += total
        bucket["on_time"] += on_time
        bucket["weighted_delay_sum"] += avg_delay * total
        bucket["weighted_crs_sum"] += crs * total
        bucket["mode_counts"][perf.mode] = bucket["mode_counts"].get(perf.mode, 0) + total

        lane_label = f"{perf.origin_region} -> {perf.destination_region}"
        lane_otd = (_to_float(perf.on_time_count, 0.0) / total) if total else 0.0
        bucket["lanes"].append({"lane": lane_label, "otd_rate": lane_otd})

        row_period = datetime(perf.period_year, perf.period_month, 1)
        if row_period >= midpoint:
            bucket["current_total"] += total
            bucket["current_on_time"] += on_time
        else:
            bucket["prior_total"] += total
            bucket["prior_on_time"] += on_time

    result = []
    for _, bucket in grouped.items():
        total_shipments = bucket["total_shipments"]
        otd_rate = (bucket["on_time"] / total_shipments) if total_shipments else 0.0
        avg_delay_hours = (
            bucket["weighted_delay_sum"] / total_shipments if total_shipments else 0.0
        )
        crs_score = (
            bucket["weighted_crs_sum"] / total_shipments if total_shipments else 0.0
        )

        prior_otd = (bucket["prior_on_time"] / bucket["prior_total"]) if bucket["prior_total"] else None
        current_otd = (bucket["current_on_time"] / bucket["current_total"]) if bucket["current_total"] else None
        trend = "neutral"
        if prior_otd is not None and current_otd is not None:
            delta = current_otd - prior_otd
            if delta > 0.01:
                trend = "up"
            elif delta < -0.01:
                trend = "down"

        lanes = sorted(bucket["lanes"], key=lambda item: item["otd_rate"], reverse=True)
        best_lane = lanes[0]["lane"] if lanes else "-"
        worst_lane = lanes[-1]["lane"] if lanes else "-"

        dominant_mode = "multimodal"
        if bucket["mode_counts"]:
            dominant_mode = max(bucket["mode_counts"], key=bucket["mode_counts"].get)

        result.append(
            {
                "carrier_id": bucket["carrier_id"],
                "carrier_name": bucket["carrier_name"],
                "mode": dominant_mode,
                "otd_rate": round(otd_rate, 4),
                "avg_delay_hours": round(avg_delay_hours, 2),
                "reliability_score": round(crs_score, 2),
                "total_shipments": total_shipments,
                "trend": trend,
                "best_lane": best_lane,
                "worst_lane": worst_lane,
            }
        )

    result.sort(key=lambda item: item["otd_rate"], reverse=True)
    return result


def _generate_lane_risk_data(organisation_id, start_date, end_date, db_session):
    """Aggregate lane-level risk and performance metrics."""

    shipments = (
        db_session.query(Shipment)
        .filter(
            Shipment.organisation_id == organisation_id,
            Shipment.estimated_departure >= start_date,
            Shipment.estimated_departure <= end_date,
            Shipment.is_archived.is_(False),
        )
        .all()
    )

    grouped = {}
    for shipment in shipments:
        key = (
            shipment.origin_port_code,
            shipment.destination_port_code,
            shipment.mode,
        )
        bucket = grouped.setdefault(
            key,
            {
                "origin_port_code": shipment.origin_port_code,
                "destination_port_code": shipment.destination_port_code,
                "mode": shipment.mode,
                "shipment_count": 0,
                "drs_sum": 0.0,
                "disruption_count": 0,
                "cargo_value_sum": 0.0,
                "completed_count": 0,
                "on_time_count": 0,
                "delay_hours": [],
            },
        )

        bucket["shipment_count"] += 1
        drs = _to_float(shipment.disruption_risk_score, 0.0)
        bucket["drs_sum"] += drs
        if drs > 60:
            bucket["disruption_count"] += 1
        bucket["cargo_value_sum"] += _to_float(shipment.cargo_value_inr, 0.0)

        if shipment.actual_arrival and shipment.estimated_arrival:
            bucket["completed_count"] += 1
            if shipment.actual_arrival <= shipment.estimated_arrival:
                bucket["on_time_count"] += 1
            else:
                bucket["delay_hours"].append(
                    max((shipment.actual_arrival - shipment.estimated_arrival).total_seconds() / 3600.0, 0.0)
                )

    lanes = []
    for _, bucket in grouped.items():
        total = bucket["shipment_count"]
        completed = bucket["completed_count"]
        avg_drs = bucket["drs_sum"] / total if total else 0.0
        avg_delay = mean(bucket["delay_hours"]) if bucket["delay_hours"] else 0.0
        disruption_frequency = (bucket["disruption_count"] / total) if total else 0.0
        on_time_rate = (bucket["on_time_count"] / completed) if completed else 0.0

        lanes.append(
            {
                "origin_port_code": bucket["origin_port_code"],
                "destination_port_code": bucket["destination_port_code"],
                "mode": bucket["mode"],
                "average_drs": round(avg_drs, 2),
                "average_delay_hours": round(avg_delay, 2),
                "disruption_frequency": round(disruption_frequency, 4),
                "total_cargo_value_inr": round(bucket["cargo_value_sum"], 2),
                "on_time_rate": round(on_time_rate, 4),
                "shipment_count": total,
            }
        )

    lanes.sort(key=lambda item: item["average_drs"], reverse=True)
    return lanes


def _generate_disruption_audit_data(organisation_id, start_date, end_date, db_session):
    """Create shipment-level disruption trajectory audit data."""

    rows = (
        db_session.query(Shipment, DisruptionScore)
        .join(DisruptionScore, DisruptionScore.shipment_id == Shipment.id)
        .filter(
            Shipment.organisation_id == organisation_id,
            DisruptionScore.computed_at >= start_date,
            DisruptionScore.computed_at <= end_date,
        )
        .order_by(Shipment.id.asc(), DisruptionScore.computed_at.asc())
        .all()
    )

    grouped = {}
    for shipment, score in rows:
        bucket = grouped.setdefault(
            shipment.id,
            {
                "shipment": shipment,
                "scores": [],
            },
        )
        bucket["scores"].append(score)

    audit_data = []
    for _, bucket in grouped.items():
        shipment = bucket["shipment"]
        scores = bucket["scores"]
        if not scores:
            continue

        peak = max(scores, key=lambda item: _to_float(item.drs_total, 0.0))

        warning_crossings = 0
        critical_crossings = 0
        previous = None
        for score in scores:
            current_drs = _to_float(score.drs_total, 0.0)
            if previous is not None:
                if previous < 61 <= current_drs:
                    warning_crossings += 1
                if previous < 81 <= current_drs:
                    critical_crossings += 1
            previous = current_drs

        recommendations = (
            db_session.query(RouteRecommendation)
            .filter(
                RouteRecommendation.shipment_id == shipment.id,
                RouteRecommendation.created_at >= start_date,
                RouteRecommendation.created_at <= end_date,
            )
            .all()
        )

        route_recommendation_generated = len(recommendations) > 0
        route_recommendation_approved = any(item.status == "approved" for item in recommendations)

        final_outcome = _outcome_label(shipment)

        audit_data.append(
            {
                "shipment_id": str(shipment.id),
                "external_reference": shipment.external_reference or str(shipment.id),
                "peak_drs": round(_to_float(peak.drs_total, 0.0), 2),
                "peak_drs_at": peak.computed_at,
                "triggering_sub_scores": {
                    "tvs": round(_to_float(peak.tvs, 0.0), 2),
                    "mcs": round(_to_float(peak.mcs, 0.0), 2),
                    "ehs": round(_to_float(peak.ehs, 0.0), 2),
                    "crs": round(_to_float(peak.crs, 0.0), 2),
                    "dtas": round(_to_float(peak.dtas, 0.0), 2),
                    "cps": round(_to_float(peak.cps, 0.0), 2),
                },
                "warning_crossings": warning_crossings,
                "critical_crossings": critical_crossings,
                "route_recommendation_generated": route_recommendation_generated,
                "route_recommendation_approved": route_recommendation_approved,
                "final_outcome": final_outcome,
            }
        )

    audit_data.sort(key=lambda item: item["peak_drs"], reverse=True)
    return audit_data


def _pdf_base_css(report_title: str, organisation_name: str, generated_at: str) -> str:
    return f"""
        @page {{
            size: A4;
            margin: 20mm 14mm 18mm 14mm;
        }}
        body {{
            font-family: Arial, Helvetica, sans-serif;
            color: #1A1A2E;
            font-size: 12px;
            line-height: 1.45;
        }}
        .cw-footer {{
            position: fixed;
            bottom: -8mm;
            left: 0;
            right: 0;
            border-top: 1px solid #DCE5F3;
            color: #5A6275;
            font-size: 10px;
            padding-top: 4px;
        }}
        .cw-footer-left {{ float: left; }}
        .cw-footer-right {{ float: right; }}
        .cw-footer-center {{ text-align: center; }}
        .cw-footer-center::after {{ content: "Page " counter(page) " of " counter(pages); }}
        .page-break {{ page-break-before: always; }}
        .cover {{
            min-height: 240mm;
            position: relative;
            padding-top: 20mm;
        }}
        .watermark {{
            position: absolute;
            right: 8mm;
            top: 105mm;
            transform: rotate(-20deg);
            font-size: 48px;
            color: #D6DBE5;
            opacity: 0.35;
            font-weight: 700;
        }}
        h1 {{ color: #1B3A6B; margin: 0 0 8px 0; font-size: 30px; }}
        h2 {{ color: #1B3A6B; margin: 0 0 12px 0; font-size: 20px; }}
        h3 {{ color: #1B3A6B; margin: 16px 0 8px 0; font-size: 14px; }}
        .muted {{ color: #6B7280; }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr 1fr 1fr 1fr;
            gap: 8px;
            margin: 10px 0 16px 0;
        }}
        .kpi-box {{
            border: 1px solid #DCE5F3;
            border-radius: 6px;
            padding: 10px;
            background: #F8FBFF;
        }}
        .kpi-value {{
            font-size: 17px;
            font-weight: 700;
            color: #1B3A6B;
            margin-bottom: 2px;
        }}
        .kpi-label {{
            font-size: 10px;
            color: #5A6275;
            text-transform: uppercase;
            letter-spacing: 0.04em;
        }}
        .kpi-trend {{ font-size: 10px; margin-top: 4px; }}
        .trend-up {{ color: #00A86B; }}
        .trend-down {{ color: #D32F2F; }}
        .trend-neutral {{ color: #6B7280; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 14px;
            table-layout: fixed;
            word-wrap: break-word;
        }}
        th {{
            background: #1B3A6B;
            color: #FFFFFF;
            font-size: 10px;
            text-transform: uppercase;
            letter-spacing: 0.03em;
            border: 1px solid #E2E8F0;
            padding: 7px;
            text-align: left;
        }}
        td {{
            border: 1px solid #E2E8F0;
            padding: 7px;
            font-size: 11px;
            vertical-align: top;
        }}
        tr:nth-child(even) td {{ background: #F7FAFF; }}
        .badge {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 999px;
            font-size: 10px;
            font-weight: 700;
        }}
        .badge-green {{ background: #E8F8F1; color: #007A4B; }}
        .badge-watch {{ background: #FFF5E5; color: #A65D00; }}
        .badge-warning {{ background: #FFF2E8; color: #B24B00; }}
        .badge-critical {{ background: #FDEDEC; color: #A82424; }}
        .cover-line {{ margin: 4px 0; color: #4B5563; }}
        .logo-text {{
            font-size: 24px;
            font-weight: 800;
            color: #1B3A6B;
            letter-spacing: 0.02em;
            margin-bottom: 16px;
        }}
        .section-note {{
            background: #EBF4FF;
            border-left: 3px solid #0077CC;
            padding: 8px 10px;
            margin-bottom: 12px;
            color: #334155;
        }}
    """


def _build_pdf_html(report_title, organisation, start_date, end_date, generated_at, body_html):
    css = _pdf_base_css(report_title, organisation.name, generated_at)
    template = """
    <html>
      <head><style>{{ css }}</style></head>
      <body>
        <div class="cw-footer">
          <div class="cw-footer-left">{{ report_title }} | {{ organisation.name }}</div>
          <div class="cw-footer-center"></div>
          <div class="cw-footer-right">Generated {{ generated_at }} | Generated by ChainWatch Pro</div>
        </div>
        {{ body_html | safe }}
      </body>
    </html>
    """
    return render_template_string(
        template,
        css=css,
        report_title=report_title,
        organisation=organisation,
        generated_at=generated_at,
        body_html=body_html,
    )


def generate_monthly_performance_pdf(data, organisation, start_date, end_date, app_context):
    """Generate monthly performance PDF bytes."""

    generated_at = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")

    template = """
    <div class="cover">
      <div class="watermark">Confidential</div>
      <div class="logo-text">ChainWatch Pro</div>
      <h1>Monthly Performance Report</h1>
      <p class="cover-line"><strong>Organisation:</strong> {{ organisation.name }}</p>
      <p class="cover-line"><strong>Date Range:</strong> {{ start_date.strftime('%d %b %Y') }} - {{ end_date.strftime('%d %b %Y') }}</p>
      <p class="cover-line"><strong>Generated At:</strong> {{ generated_at }}</p>
      <p class="cover-line muted">See the disruption before it sees you.</p>
    </div>

    <div class="page-break"></div>

    <h2>Executive Summary</h2>
    <div class="kpi-grid">
      <div class="kpi-box">
        <div class="kpi-value">{{ data.total_shipments }}</div>
        <div class="kpi-label">Total Shipments</div>
        {% set trend = data.kpi_trends.total_shipments %}
        <div class="kpi-trend trend-{{ trend.direction }}">{{ '%+.2f'|format(trend.delta_pct) }}% vs prior period</div>
      </div>
      <div class="kpi-box">
        <div class="kpi-value">{{ '%.2f'|format(data.otd_rate * 100) }}%</div>
        <div class="kpi-label">OTD Rate</div>
        {% set trend = data.kpi_trends.otd_rate %}
        <div class="kpi-trend trend-{{ trend.direction }}">{{ '%+.2f'|format(trend.delta_pct) }}% vs prior period</div>
      </div>
      <div class="kpi-box">
        <div class="kpi-value">{{ '%.2f'|format(data.average_drs) }}</div>
        <div class="kpi-label">Average DRS</div>
        {% set trend = data.kpi_trends.average_drs %}
        <div class="kpi-trend trend-{{ trend.direction }}">{{ '%+.2f'|format(trend.delta_pct) }}% vs prior period</div>
      </div>
      <div class="kpi-box">
        <div class="kpi-value">{{ data.critical_alerts_generated }}</div>
        <div class="kpi-label">Critical Alerts</div>
        {% set trend = data.kpi_trends.critical_alerts %}
        <div class="kpi-trend trend-{{ trend.direction }}">{{ '%+.2f'|format(trend.delta_pct) }}% vs prior period</div>
      </div>
      <div class="kpi-box">
        <div class="kpi-value">{{ savings_label }}</div>
        <div class="kpi-label">Estimated ₹ Savings</div>
        {% set trend = data.kpi_trends.estimated_savings_inr %}
        <div class="kpi-trend trend-{{ trend.direction }}">{{ '%+.2f'|format(trend.delta_pct) }}% vs prior period</div>
      </div>
    </div>

    <div class="section-note">Alerts acknowledged: {{ data.alerts_acknowledged }} | Average resolution: {{ '%.2f'|format(data.avg_alert_resolution_hours) }} hours | Value at risk: {{ value_at_risk }}</div>

    <h3>Carrier Performance</h3>
    <table>
      <thead>
        <tr><th>Carrier</th><th>Shipment Count</th><th>OTD %</th></tr>
      </thead>
      <tbody>
        {% for carrier in data.top_carriers %}
        <tr>
          <td>{{ carrier.carrier_name }}</td>
          <td>{{ carrier.shipment_count }}</td>
          <td>{{ '%.2f'|format(carrier.otd_rate * 100) }}%</td>
        </tr>
        {% else %}
        <tr><td colspan="3">No carrier performance records in this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>

    <h3>Shipment Volume by Mode</h3>
    <table>
      <thead>
        <tr><th>Mode</th><th>Shipments</th></tr>
      </thead>
      <tbody>
        {% for row in data.shipment_volume_by_mode %}
        <tr><td>{{ row.mode|replace('_', ' ')|title }}</td><td>{{ row.count }}</td></tr>
        {% else %}
        <tr><td colspan="2">No mode distribution records in this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>

    <h3>Week-by-Week OTD Trend</h3>
    <table>
      <thead>
        <tr><th>ISO Week</th><th>Completed</th><th>On-Time</th><th>OTD %</th></tr>
      </thead>
      <tbody>
        {% for row in data.week_otd_trend %}
        <tr>
          <td>{{ row.iso_week }}</td>
          <td>{{ row.completed }}</td>
          <td>{{ row.on_time }}</td>
          <td>{{ '%.2f'|format(row.otd_rate * 100) }}%</td>
        </tr>
        {% else %}
        <tr><td colspan="4">No completed shipments for week trend in this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>

    <h3>Rerouting Decisions Summary</h3>
    <table>
      <thead>
        <tr>
          <th>Date</th><th>Shipment Reference</th><th>Option</th><th>₹ Cost Delta</th><th>Revised ETA</th><th>Outcome</th>
        </tr>
      </thead>
      <tbody>
        {% for row in data.rerouting_decisions %}
        <tr>
          <td>{{ row.date.strftime('%d %b %Y') if row.date else '-' }}</td>
          <td>{{ row.shipment_reference }}</td>
          <td>{{ row.option_chosen }}</td>
          <td>{{ '%.2f'|format(row.cost_delta_inr) }}</td>
          <td>{{ row.revised_eta.strftime('%d %b %Y %H:%M') if row.revised_eta else '-' }}</td>
          <td>{{ row.outcome }}</td>
        </tr>
        {% else %}
        <tr><td colspan="6">No rerouting decisions recorded in this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>
    """

    body_html = render_template_string(
        template,
        data=data,
        organisation=organisation,
        start_date=start_date,
        end_date=end_date,
        generated_at=generated_at,
        savings_label=_format_inr_lakh_crore(data.get("estimated_savings_inr", 0.0)),
        value_at_risk=_format_inr_lakh_crore(data.get("cargo_value_at_risk_inr", 0.0)),
    )

    html_string = _build_pdf_html(
        "Monthly Performance Report",
        organisation,
        start_date,
        end_date,
        generated_at,
        body_html,
    )
    return HTML(string=html_string).write_pdf()


def generate_carrier_comparison_pdf(data, organisation, start_date, end_date, app_context):
    """Generate carrier comparison PDF bytes."""

    generated_at = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")

    template = """
    <div class="cover">
      <div class="watermark">Confidential</div>
      <div class="logo-text">ChainWatch Pro</div>
      <h1>Carrier Comparison Report</h1>
      <p class="cover-line"><strong>Organisation:</strong> {{ organisation.name }}</p>
      <p class="cover-line"><strong>Date Range:</strong> {{ start_date.strftime('%d %b %Y') }} - {{ end_date.strftime('%d %b %Y') }}</p>
      <p class="cover-line"><strong>Generated At:</strong> {{ generated_at }}</p>
    </div>

    <div class="page-break"></div>

    <h2>Carrier Benchmark Table</h2>
    <table>
      <thead>
        <tr>
          <th>Carrier</th>
          <th>Mode</th>
          <th>OTD %</th>
          <th>Avg Delay (hrs)</th>
          <th>CRS</th>
          <th>Total Shipments</th>
          <th>Trend</th>
          <th>Best Lane</th>
          <th>Worst Lane</th>
        </tr>
      </thead>
      <tbody>
        {% for row in data %}
        <tr>
          <td>{{ row.carrier_name }}</td>
          <td>{{ row.mode|replace('_', ' ')|title }}</td>
          <td>{{ '%.2f'|format(row.otd_rate * 100) }}%</td>
          <td>{{ '%.2f'|format(row.avg_delay_hours) }}</td>
          <td>{{ '%.2f'|format(row.reliability_score) }}</td>
          <td>{{ row.total_shipments }}</td>
          <td>{{ row.trend|title }}</td>
          <td>{{ row.best_lane }}</td>
          <td>{{ row.worst_lane }}</td>
        </tr>
        {% else %}
        <tr><td colspan="9">No carrier benchmark rows found for this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>
    """

    body_html = render_template_string(
        template,
        data=data,
        organisation=organisation,
        start_date=start_date,
        end_date=end_date,
        generated_at=generated_at,
    )

    html_string = _build_pdf_html(
        "Carrier Comparison Report",
        organisation,
        start_date,
        end_date,
        generated_at,
        body_html,
    )
    return HTML(string=html_string).write_pdf()


def generate_lane_risk_pdf(data, organisation, start_date, end_date, app_context):
    """Generate lane risk analysis PDF bytes."""

    generated_at = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")

    template = """
    <div class="cover">
      <div class="watermark">Confidential</div>
      <div class="logo-text">ChainWatch Pro</div>
      <h1>Lane Risk Analysis Report</h1>
      <p class="cover-line"><strong>Organisation:</strong> {{ organisation.name }}</p>
      <p class="cover-line"><strong>Date Range:</strong> {{ start_date.strftime('%d %b %Y') }} - {{ end_date.strftime('%d %b %Y') }}</p>
      <p class="cover-line"><strong>Generated At:</strong> {{ generated_at }}</p>
    </div>

    <div class="page-break"></div>

    <h2>Highest Risk Lanes</h2>
    <table>
      <thead>
        <tr>
          <th>Lane</th><th>Mode</th><th>Avg DRS</th><th>Avg Delay (hrs)</th><th>Disruption Frequency</th><th>Cargo Value (₹)</th><th>On-Time %</th><th>Shipments</th>
        </tr>
      </thead>
      <tbody>
        {% for row in data %}
        <tr>
          <td>{{ row.origin_port_code }} → {{ row.destination_port_code }}</td>
          <td>{{ row.mode|replace('_', ' ')|title }}</td>
          <td>{{ '%.2f'|format(row.average_drs) }}</td>
          <td>{{ '%.2f'|format(row.average_delay_hours) }}</td>
          <td>{{ '%.2f'|format(row.disruption_frequency * 100) }}%</td>
          <td>{{ '%.2f'|format(row.total_cargo_value_inr) }}</td>
          <td>{{ '%.2f'|format(row.on_time_rate * 100) }}%</td>
          <td>{{ row.shipment_count }}</td>
        </tr>
        {% else %}
        <tr><td colspan="8">No lane-risk rows found for this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>
    """

    body_html = render_template_string(
        template,
        data=data,
        organisation=organisation,
        start_date=start_date,
        end_date=end_date,
        generated_at=generated_at,
    )

    html_string = _build_pdf_html(
        "Lane Risk Analysis Report",
        organisation,
        start_date,
        end_date,
        generated_at,
        body_html,
    )
    return HTML(string=html_string).write_pdf()


def generate_disruption_audit_pdf(data, organisation, start_date, end_date, app_context):
    """Generate disruption audit PDF bytes."""

    generated_at = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")

    template = """
    <div class="cover">
      <div class="watermark">Confidential</div>
      <div class="logo-text">ChainWatch Pro</div>
      <h1>Disruption Audit Report</h1>
      <p class="cover-line"><strong>Organisation:</strong> {{ organisation.name }}</p>
      <p class="cover-line"><strong>Date Range:</strong> {{ start_date.strftime('%d %b %Y') }} - {{ end_date.strftime('%d %b %Y') }}</p>
      <p class="cover-line"><strong>Generated At:</strong> {{ generated_at }}</p>
    </div>

    <div class="page-break"></div>

    <h2>Shipment Disruption Audit</h2>
    <table>
      <thead>
        <tr>
          <th>Shipment</th><th>Peak DRS</th><th>Peak Date</th><th>Warning Crossings</th><th>Critical Crossings</th><th>Recommendation</th><th>Approved</th><th>Final Outcome</th>
        </tr>
      </thead>
      <tbody>
        {% for row in data %}
        <tr>
          <td>{{ row.external_reference }}</td>
          <td>{{ '%.2f'|format(row.peak_drs) }}</td>
          <td>{{ row.peak_drs_at.strftime('%d %b %Y %H:%M') if row.peak_drs_at else '-' }}</td>
          <td>{{ row.warning_crossings }}</td>
          <td>{{ row.critical_crossings }}</td>
          <td>{{ 'Yes' if row.route_recommendation_generated else 'No' }}</td>
          <td>{{ 'Yes' if row.route_recommendation_approved else 'No' }}</td>
          <td>{{ row.final_outcome }}</td>
        </tr>
        {% else %}
        <tr><td colspan="8">No disruption audit rows found for this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>

    <h3>Triggering Sub-Scores at Peak</h3>
    <table>
      <thead>
        <tr><th>Shipment</th><th>TVS</th><th>MCS</th><th>EHS</th><th>CRS</th><th>DTAS</th><th>CPS</th></tr>
      </thead>
      <tbody>
        {% for row in data %}
        <tr>
          <td>{{ row.external_reference }}</td>
          <td>{{ row.triggering_sub_scores.tvs }}</td>
          <td>{{ row.triggering_sub_scores.mcs }}</td>
          <td>{{ row.triggering_sub_scores.ehs }}</td>
          <td>{{ row.triggering_sub_scores.crs }}</td>
          <td>{{ row.triggering_sub_scores.dtas }}</td>
          <td>{{ row.triggering_sub_scores.cps }}</td>
        </tr>
        {% else %}
        <tr><td colspan="7">No peak score rows found for this period.</td></tr>
        {% endfor %}
      </tbody>
    </table>
    """

    body_html = render_template_string(
        template,
        data=data,
        organisation=organisation,
        start_date=start_date,
        end_date=end_date,
        generated_at=generated_at,
    )

    html_string = _build_pdf_html(
        "Disruption Audit Report",
        organisation,
        start_date,
        end_date,
        generated_at,
        body_html,
    )
    return HTML(string=html_string).write_pdf()


def _workbook_styles():
    navy_fill = PatternFill("solid", fgColor="1B3A6B")
    light_blue_fill = PatternFill("solid", fgColor="EBF4FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="1B3A6B", bold=True, size=16)
    label_font = Font(color="1A1A2E", bold=True)
    value_font = Font(color="1B3A6B", bold=True, size=13)
    border = Border(
        left=Side(style="thin", color="D9E2F2"),
        right=Side(style="thin", color="D9E2F2"),
        top=Side(style="thin", color="D9E2F2"),
        bottom=Side(style="thin", color="D9E2F2"),
    )
    return {
        "navy_fill": navy_fill,
        "light_blue_fill": light_blue_fill,
        "white_fill": white_fill,
        "header_font": header_font,
        "title_font": title_font,
        "label_font": label_font,
        "value_font": value_font,
        "border": border,
    }


def _style_header_row(worksheet, row_index, column_count, styles):
    for col in range(1, column_count + 1):
        cell = worksheet.cell(row=row_index, column=col)
        cell.fill = styles["navy_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["border"]


def _auto_width(worksheet, max_width=42):
    for column_cells in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[col_letter].width = min(max(max_length + 2, 10), max_width)


def _apply_table_borders_and_striping(worksheet, start_row, end_row, column_count, styles):
    for row in range(start_row, end_row + 1):
        fill = styles["light_blue_fill"] if row % 2 == 0 else styles["white_fill"]
        for col in range(1, column_count + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = styles["border"]
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _wb_to_bytes(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def generate_monthly_performance_excel(data, organisation, start_date, end_date):
    """Generate monthly performance workbook bytes."""

    workbook = Workbook()
    styles = _workbook_styles()

    summary = workbook.active
    summary.title = "Summary"
    summary.merge_cells("A1:F1")
    summary["A1"] = (
        f"Monthly Performance Report | {organisation.name} | "
        f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    )
    summary["A1"].font = styles["title_font"]

    kpi_rows = [
        ("Total Shipments", data.get("total_shipments", 0)),
        ("OTD Rate", data.get("otd_rate", 0.0)),
        ("Average DRS", data.get("average_drs", 0.0)),
        ("Critical Alerts", data.get("critical_alerts_generated", 0)),
        ("Estimated ₹ Savings", data.get("estimated_savings_inr", 0.0)),
    ]

    row_index = 3
    for label, value in kpi_rows:
        summary.cell(row=row_index, column=1, value=label).font = styles["label_font"]
        summary.cell(row=row_index, column=2, value=value).font = styles["value_font"]
        summary.cell(row=row_index, column=1).fill = styles["light_blue_fill"]
        summary.cell(row=row_index, column=2).fill = styles["white_fill"]
        summary.cell(row=row_index, column=1).border = styles["border"]
        summary.cell(row=row_index, column=2).border = styles["border"]
        row_index += 1

    summary["B4"].number_format = "0.00%"
    summary["B7"].number_format = "₹#,##,##0.00"
    _auto_width(summary)

    shipment_ws = workbook.create_sheet("Shipment Detail")
    shipment_headers = [
        "Shipment ID",
        "External Reference",
        "Carrier",
        "Mode",
        "Origin",
        "Destination",
        "Status",
        "Estimated Departure",
        "Estimated Arrival",
        "Actual Arrival",
        "Delay Hours",
        "Peak DRS",
        "Cargo Value (₹)",
    ]
    shipment_ws.append(shipment_headers)
    _style_header_row(shipment_ws, 1, len(shipment_headers), styles)

    for row in data.get("shipment_detail", []):
        shipment_ws.append(
            [
                row.get("shipment_id"),
                row.get("external_reference"),
                row.get("carrier"),
                row.get("mode"),
                row.get("origin"),
                row.get("destination"),
                row.get("status"),
                row.get("estimated_departure"),
                row.get("estimated_arrival"),
                row.get("actual_arrival"),
                row.get("delay_hours"),
                row.get("peak_drs"),
                row.get("cargo_value_inr"),
            ]
        )

    if shipment_ws.max_row > 1:
        _apply_table_borders_and_striping(shipment_ws, 2, shipment_ws.max_row, len(shipment_headers), styles)
    shipment_ws.auto_filter.ref = f"A1:{get_column_letter(len(shipment_headers))}{shipment_ws.max_row}"

    for row in range(2, shipment_ws.max_row + 1):
        shipment_ws.cell(row=row, column=8).number_format = "DD-MMM-YYYY HH:MM"
        shipment_ws.cell(row=row, column=9).number_format = "DD-MMM-YYYY HH:MM"
        shipment_ws.cell(row=row, column=10).number_format = "DD-MMM-YYYY HH:MM"
        shipment_ws.cell(row=row, column=13).number_format = "₹#,##,##0.00"

    _auto_width(shipment_ws)

    carrier_ws = workbook.create_sheet("Carrier Performance")
    carrier_headers = [
        "Carrier Name",
        "Mode",
        "OTD%",
        "Avg Delay Hours",
        "Total Shipments",
        "CRS Score",
    ]
    carrier_ws.append(carrier_headers)
    _style_header_row(carrier_ws, 1, len(carrier_headers), styles)

    for row in data.get("carrier_performance", []):
        carrier_ws.append(
            [
                row.get("carrier_name"),
                row.get("mode"),
                row.get("otd_rate"),
                row.get("avg_delay_hours"),
                row.get("total_shipments"),
                row.get("crs_score"),
            ]
        )

    if carrier_ws.max_row > 1:
        _apply_table_borders_and_striping(carrier_ws, 2, carrier_ws.max_row, len(carrier_headers), styles)
    carrier_ws.auto_filter.ref = f"A1:{get_column_letter(len(carrier_headers))}{carrier_ws.max_row}"
    for row in range(2, carrier_ws.max_row + 1):
        carrier_ws.cell(row=row, column=3).number_format = "0.00%"
    _auto_width(carrier_ws)

    alerts_ws = workbook.create_sheet("Alerts")
    alert_headers = [
        "Alert Type",
        "Severity",
        "Title",
        "Shipment Reference",
        "Created At",
        "Acknowledged At",
        "Resolution Hours",
    ]
    alerts_ws.append(alert_headers)
    _style_header_row(alerts_ws, 1, len(alert_headers), styles)

    for row in data.get("alerts", []):
        alerts_ws.append(
            [
                row.get("alert_type"),
                row.get("severity"),
                row.get("title"),
                row.get("shipment_reference"),
                row.get("created_at"),
                row.get("acknowledged_at"),
                row.get("resolution_hours"),
            ]
        )

    if alerts_ws.max_row > 1:
        _apply_table_borders_and_striping(alerts_ws, 2, alerts_ws.max_row, len(alert_headers), styles)
    alerts_ws.auto_filter.ref = f"A1:{get_column_letter(len(alert_headers))}{alerts_ws.max_row}"
    for row in range(2, alerts_ws.max_row + 1):
        alerts_ws.cell(row=row, column=5).number_format = "DD-MMM-YYYY HH:MM"
        alerts_ws.cell(row=row, column=6).number_format = "DD-MMM-YYYY HH:MM"
    _auto_width(alerts_ws)

    return _wb_to_bytes(workbook)


def generate_carrier_comparison_excel(data, organisation, start_date, end_date):
    """Generate carrier comparison workbook bytes."""

    workbook = Workbook()
    styles = _workbook_styles()

    ws = workbook.active
    ws.title = "Carrier Comparison"

    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"Carrier Comparison Report | {organisation.name} | "
        f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    )
    ws["A1"].font = styles["title_font"]

    headers = [
        "Carrier",
        "Mode",
        "OTD%",
        "Avg Delay Hours",
        "CRS Score",
        "Total Shipments",
        "Trend",
        "Best Lane",
        "Worst Lane",
    ]
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 3, len(headers), styles)

    for row in data:
        ws.append(
            [
                row.get("carrier_name"),
                row.get("mode"),
                row.get("otd_rate"),
                row.get("avg_delay_hours"),
                row.get("reliability_score"),
                row.get("total_shipments"),
                row.get("trend"),
                row.get("best_lane"),
                row.get("worst_lane"),
            ]
        )

    if ws.max_row > 3:
        _apply_table_borders_and_striping(ws, 4, ws.max_row, len(headers), styles)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"
    for row_idx in range(4, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "0.00%"

    _auto_width(ws)
    return _wb_to_bytes(workbook)


def generate_lane_risk_excel(data, organisation, start_date, end_date):
    """Generate lane risk workbook bytes."""

    workbook = Workbook()
    styles = _workbook_styles()

    ws = workbook.active
    ws.title = "Lane Risk"

    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"Lane Risk Analysis Report | {organisation.name} | "
        f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    )
    ws["A1"].font = styles["title_font"]

    headers = [
        "Origin",
        "Destination",
        "Mode",
        "Avg DRS",
        "Avg Delay Hours",
        "Disruption Frequency",
        "Total Cargo Value (₹)",
        "On-Time Rate",
        "Shipments",
    ]

    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 3, len(headers), styles)

    for row in data:
        ws.append(
            [
                row.get("origin_port_code"),
                row.get("destination_port_code"),
                row.get("mode"),
                row.get("average_drs"),
                row.get("average_delay_hours"),
                row.get("disruption_frequency"),
                row.get("total_cargo_value_inr"),
                row.get("on_time_rate"),
                row.get("shipment_count"),
            ]
        )

    if ws.max_row > 3:
        _apply_table_borders_and_striping(ws, 4, ws.max_row, len(headers), styles)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"

    for row_idx in range(4, ws.max_row + 1):
        ws.cell(row=row_idx, column=6).number_format = "0.00%"
        ws.cell(row=row_idx, column=7).number_format = "₹#,##,##0.00"
        ws.cell(row=row_idx, column=8).number_format = "0.00%"

    _auto_width(ws)
    return _wb_to_bytes(workbook)


def generate_disruption_audit_excel(data, organisation, start_date, end_date):
    """Generate disruption audit workbook bytes."""

    workbook = Workbook()
    styles = _workbook_styles()

    summary = workbook.active
    summary.title = "Disruption Audit"

    summary.merge_cells("A1:H1")
    summary["A1"] = (
        f"Disruption Audit Report | {organisation.name} | "
        f"{start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    )
    summary["A1"].font = styles["title_font"]

    headers = [
        "Shipment",
        "Peak DRS",
        "Peak DRS At",
        "Warning Crossings",
        "Critical Crossings",
        "Recommendation Generated",
        "Recommendation Approved",
        "Final Outcome",
    ]
    summary.append([])
    summary.append(headers)
    _style_header_row(summary, 3, len(headers), styles)

    for row in data:
        summary.append(
            [
                row.get("external_reference"),
                row.get("peak_drs"),
                row.get("peak_drs_at"),
                row.get("warning_crossings"),
                row.get("critical_crossings"),
                row.get("route_recommendation_generated"),
                row.get("route_recommendation_approved"),
                row.get("final_outcome"),
            ]
        )

    if summary.max_row > 3:
        _apply_table_borders_and_striping(summary, 4, summary.max_row, len(headers), styles)
    summary.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{summary.max_row}"
    for row_idx in range(4, summary.max_row + 1):
        summary.cell(row=row_idx, column=3).number_format = "DD-MMM-YYYY HH:MM"

    _auto_width(summary)

    trigger_ws = workbook.create_sheet("Peak Sub Scores")
    trigger_headers = [
        "Shipment",
        "TVS",
        "MCS",
        "EHS",
        "CRS",
        "DTAS",
        "CPS",
    ]
    trigger_ws.append(trigger_headers)
    _style_header_row(trigger_ws, 1, len(trigger_headers), styles)

    for row in data:
        subs = row.get("triggering_sub_scores") or {}
        trigger_ws.append(
            [
                row.get("external_reference"),
                subs.get("tvs"),
                subs.get("mcs"),
                subs.get("ehs"),
                subs.get("crs"),
                subs.get("dtas"),
                subs.get("cps"),
            ]
        )

    if trigger_ws.max_row > 1:
        _apply_table_borders_and_striping(trigger_ws, 2, trigger_ws.max_row, len(trigger_headers), styles)
    trigger_ws.auto_filter.ref = f"A1:{get_column_letter(len(trigger_headers))}{trigger_ws.max_row}"
    _auto_width(trigger_ws)

    return _wb_to_bytes(workbook)


__all__ = [
    "REPORT_TYPES",
    "generate_report",
    "_generate_monthly_performance_data",
    "_generate_carrier_comparison_data",
    "_generate_lane_risk_data",
    "_generate_disruption_audit_data",
    "generate_monthly_performance_pdf",
    "generate_monthly_performance_excel",
    "generate_carrier_comparison_pdf",
    "generate_carrier_comparison_excel",
    "generate_lane_risk_pdf",
    "generate_lane_risk_excel",
    "generate_disruption_audit_pdf",
    "generate_disruption_audit_excel",
]
